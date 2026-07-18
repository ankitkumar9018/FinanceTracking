"""Tests for the new import/export data formats.

Covered without network or external files:
- ``parse_ofx`` on an inline OFX investment statement.
- ``parse_qif`` on an inline ``!Type:Invest`` statement.
- ``import_statement`` round-trips OFX rows into holdings/transactions.
- ``export_workbook_xlsx`` produces a re-openable multi-sheet workbook.
- ``export_everything_zip`` produces a valid ZIP with the expected members.
- ``parse_cas`` raises a clear RuntimeError when ``casparser`` is absent
  (or is skipped when the optional package is installed).

Run with:
    uv run pytest tests/test_import_formats.py -q
"""

from __future__ import annotations

import importlib.util
import io
import zipfile
from datetime import date

import pytest
from openpyxl import load_workbook
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.dividend import Dividend
from app.models.holding import Holding
from app.models.portfolio import Portfolio
from app.models.transaction import Transaction
from app.models.user import User
from app.services.cas_import_service import parse_cas
from app.services.export_service import export_everything_zip, export_workbook_xlsx
from app.services.ofx_qif_import_service import import_statement, parse_ofx, parse_qif

# ---------------------------------------------------------------------------
# Inline fixtures
# ---------------------------------------------------------------------------

OFX_INVESTMENT = """OFXHEADER:100
DATA:OFXSGML
VERSION:102

<OFX>
<INVSTMTMSGSRSV1>
<INVSTMTTRNRS>
<INVSTMTRS>
<INVTRANLIST>
<BUYSTOCK>
<INVBUY>
<INVTRAN>
<FITID>1001
<DTTRADE>20240115120000.000[-5:EST]
</INVTRAN>
<SECID>
<UNIQUEID>INE002A01018
<UNIQUEIDTYPE>ISIN
</SECID>
<UNITS>10
<UNITPRICE>2500.00
<COMMISSION>50.00
<TOTAL>-25050.00
</INVBUY>
<BUYTYPE>BUY
</BUYSTOCK>
</INVTRANLIST>
</INVSTMTRS>
</INVSTMTTRNRS>
</INVSTMTMSGSRSV1>
<SECLISTMSGSRSV1>
<SECLIST>
<STOCKINFO>
<SECINFO>
<SECID>
<UNIQUEID>INE002A01018
<UNIQUEIDTYPE>ISIN
</SECID>
<SECNAME>Reliance Industries
<TICKER>RELIANCE
</SECINFO>
</STOCKINFO>
</SECLIST>
</SECLISTMSGSRSV1>
</OFX>
"""

QIF_INVESTMENT = """!Type:Invest
D01/15'24
NBuy
YReliance Industries
I2500.00
Q10
O50.00
T25050.00
^
D02/20'24
NSell
YReliance Industries
I2800.00
Q5
T13950.00
^
"""


# ---------------------------------------------------------------------------
# DB seeding helper
# ---------------------------------------------------------------------------

async def _seed_portfolio(db: AsyncSession) -> Portfolio:
    """Create a user + portfolio with two priced holdings, a txn, a dividend."""
    user = User(email="fmt@example.com", password_hash="x", display_name="Fmt Tester")
    db.add(user)
    await db.flush()

    portfolio = Portfolio(user_id=user.id, name="Formats Portfolio", currency="INR")
    db.add(portfolio)
    await db.flush()

    h1 = Holding(
        portfolio_id=portfolio.id,
        stock_symbol="RELIANCE",
        stock_name="Reliance Industries",
        exchange="NSE",
        currency="INR",
        cumulative_quantity=10.0,
        average_price=2500.0,
        current_price=2800.0,
    )
    h2 = Holding(
        portfolio_id=portfolio.id,
        stock_symbol="TCS",
        stock_name="Tata Consultancy",
        exchange="NSE",
        currency="INR",
        cumulative_quantity=5.0,
        average_price=3500.0,
        current_price=3400.0,
    )
    db.add_all([h1, h2])
    await db.flush()

    db.add(Transaction(
        holding_id=h1.id, transaction_type="BUY", date=date(2024, 1, 15),
        quantity=10.0, price=2500.0, brokerage=50.0,
    ))
    db.add(Dividend(
        holding_id=h1.id, ex_date=date(2024, 6, 15), payment_date=date(2024, 7, 1),
        amount_per_share=10.5, total_amount=105.0, is_reinvested=False,
    ))
    await db.flush()
    return portfolio


# ---------------------------------------------------------------------------
# OFX / QIF parsers
# ---------------------------------------------------------------------------

def test_parse_ofx_investment_yields_buy():
    rows = parse_ofx(OFX_INVESTMENT.encode("utf-8"))
    assert len(rows) >= 1
    buys = [r for r in rows if r["transaction_type"] == "BUY"]
    assert buys, "expected at least one BUY row"
    r = buys[0]
    assert r["stock_symbol"] == "RELIANCE"  # resolved via TICKER from SECINFO
    assert r["quantity"] == 10.0
    assert r["price"] == 2500.0
    assert r["brokerage"] == 50.0
    assert r["date"] == date(2024, 1, 15)
    assert r["exchange"]  # non-empty (required by import_to_portfolio)


def test_parse_qif_investment_yields_rows():
    rows = parse_qif(QIF_INVESTMENT.encode("utf-8"))
    assert len(rows) >= 1
    types = {r["transaction_type"] for r in rows}
    assert "BUY" in types
    buy = next(r for r in rows if r["transaction_type"] == "BUY")
    assert buy["quantity"] == 10.0
    assert buy["price"] == 2500.0
    assert buy["date"] == date(2024, 1, 15)
    assert buy["stock_name"] == "Reliance Industries"


async def test_import_statement_creates_holdings(db: AsyncSession):
    """OFX rows must reconcile with the shared import_to_portfolio shape."""
    user = User(email="ofximp@example.com", password_hash="x", display_name="OFX")
    db.add(user)
    await db.flush()
    portfolio = Portfolio(user_id=user.id, name="OFX Portfolio", currency="INR")
    db.add(portfolio)
    await db.flush()

    rows = parse_ofx(OFX_INVESTMENT.encode("utf-8"))
    summary = await import_statement(rows, portfolio.id, db)
    assert summary["transactions_created"] >= 1
    assert summary["holdings_created"] >= 1


# ---------------------------------------------------------------------------
# XLSX workbook export
# ---------------------------------------------------------------------------

async def test_export_workbook_xlsx(db: AsyncSession):
    portfolio = await _seed_portfolio(db)
    data = await export_workbook_xlsx(portfolio.id, db)
    assert isinstance(data, bytes) and len(data) > 0

    wb = load_workbook(io.BytesIO(data))
    assert wb.sheetnames == ["Holdings", "Transactions", "Dividends", "Summary"]
    # Holdings sheet: header + 2 data rows
    assert wb["Holdings"].max_row == 3
    # Dividends sheet: header + 1 data row
    assert wb["Dividends"].max_row == 2


async def test_export_workbook_xlsx_no_holdings_raises(db: AsyncSession):
    user = User(email="empty@example.com", password_hash="x", display_name="Empty")
    db.add(user)
    await db.flush()
    portfolio = Portfolio(user_id=user.id, name="Empty", currency="INR")
    db.add(portfolio)
    await db.flush()

    with pytest.raises(ValueError):
        await export_workbook_xlsx(portfolio.id, db)


# ---------------------------------------------------------------------------
# ZIP bundle export
# ---------------------------------------------------------------------------

async def test_export_everything_zip(db: AsyncSession):
    portfolio = await _seed_portfolio(db)
    data = await export_everything_zip(portfolio.id, "Fmt Tester", db)
    assert isinstance(data, bytes) and len(data) > 0

    with zipfile.ZipFile(io.BytesIO(data)) as zf:
        names = set(zf.namelist())
        for required in (
            "holdings.csv",
            "transactions.csv",
            "portfolio_backup.json",
            "report.html",
            "portfolio_workbook.xlsx",
            "README.txt",
        ):
            assert required in names, f"missing {required} in bundle"
        # Bundle contents must be non-empty / valid.
        assert zf.read("holdings.csv")
        assert zf.read("portfolio_backup.json")


# ---------------------------------------------------------------------------
# CAS import (graceful degradation)
# ---------------------------------------------------------------------------

_CASPARSER_INSTALLED = importlib.util.find_spec("casparser") is not None


@pytest.mark.skipif(
    _CASPARSER_INSTALLED,
    reason="casparser installed — the RuntimeError degradation path is not exercised",
)
def test_parse_cas_raises_without_casparser():
    with pytest.raises(RuntimeError, match="casparser"):
        parse_cas(b"%PDF-1.4 fake", password=None)


@pytest.mark.skipif(
    not _CASPARSER_INSTALLED,
    reason="requires the optional 'cas' extra (casparser) to validate the mapping",
)
def test_cas_mapping_matches_casparser_schema():
    """Prove ``_map_cas_to_mf`` against the EXACT dict casparser emits.

    We can't decrypt a real CAS PDF in CI, but we can build the dict casparser's
    ``read_cas_pdf(..., output="dict")`` produces — straight from casparser's own
    pydantic models — and confirm the never-network mapping logic is correct
    (amfi→code with isin fallback, close→units, valuation nav/cost, and the
    cost→value invested fallback), including Decimal→float coercion.
    """
    from decimal import Decimal

    from casparser.enums import CASFileType, FileType
    from casparser.types import (
        CASData,
        Folio,
        InvestorInfo,
        Scheme,
        SchemeValuation,
        StatementPeriod,
    )

    from app.services.cas_import_service import _map_cas_to_mf

    cas = CASData(
        statement_period=StatementPeriod(**{"from": "2024-01-01", "to": "2024-03-31"}),
        investor_info=InvestorInfo(name="A", email="a@b.c", address="x", mobile="9"),
        cas_type=CASFileType.SUMMARY,
        file_type=FileType.CAMS,
        folios=[
            Folio(
                folio="12345/67",
                amc="HDFC AMC",
                PAN="ABCDE1234F",
                schemes=[
                    Scheme(
                        scheme="HDFC Flexi Cap Fund - Growth",
                        rta_code="H123", rta="CAMS", type="EQUITY",
                        isin="INF179K01608", amfi="118989",
                        open=Decimal("0"), close=Decimal("1234.567"),
                        close_calculated=Decimal("1234.567"),
                        valuation=SchemeValuation(
                            date="2024-03-31", nav=Decimal("1450.25"),
                            cost=Decimal("1500000"), value=Decimal("1790123.45"),
                        ),
                        transactions=[],
                    ),
                    Scheme(  # amfi missing -> isin fallback; cost missing -> value
                        scheme="SBI Small Cap - Growth",
                        rta_code="S9", rta="KFINTECH", type="EQUITY",
                        isin="INF200K01594", amfi=None,
                        open=Decimal("0"), close=Decimal("500"),
                        close_calculated=Decimal("500"),
                        valuation=SchemeValuation(
                            date="2024-03-31", nav=Decimal("180.5"),
                            cost=None, value=Decimal("90250"),
                        ),
                        transactions=[],
                    ),
                ],
            )
        ],
    )
    # Exactly what read_cas_pdf(..., output="dict") returns:
    rows = _map_cas_to_mf(cas.model_dump(by_alias=True))

    assert len(rows) == 2
    r0, r1 = rows
    assert r0["scheme_code"] == "118989"
    assert r0["scheme_name"].startswith("HDFC Flexi Cap")
    assert r0["folio_number"] == "12345/67"
    assert r0["units"] == 1234.567
    assert r0["nav"] == 1450.25
    assert r0["invested_amount"] == 1500000.0
    # amfi absent -> isin as code; cost absent -> valuation.value as invested
    assert r1["scheme_code"] == "INF200K01594"
    assert r1["units"] == 500.0
    assert r1["invested_amount"] == 90250.0
