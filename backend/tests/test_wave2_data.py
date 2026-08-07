"""Data/import-export wave-2 regression tests.

Covers the verified fixes in the data & import/export service layer:

1.  QIF dates use ONE inferred day/month convention per file (no per-row
    flip-flop) and OFX/QIF numbers parse European decimal commas.
2.  Corporate split apply adjusts only shares held as of the ex-date —
    post-ex-date buys are unaffected — and the stored quantity/average
    derive from the transaction ledger.
3.  Dividend CSV re-import skips rows already present (holding + ex_date +
    total_amount) instead of double-counting.
4.  Backup restore dedups user-level goals/assets/tax records by natural key
    so a (double-)restore does not multiply them.
5.  Unpriced holdings render as "—" (not a -100% loss) in the HTML report,
    XLSX workbook, and Sheets CSV, and are excluded from totals.
6.  Import provenance: the CSV route stamps source="CSV"; QIF statements
    stamp source="QIF".
8.  ESG weighted averages skip missing sub-scores instead of zero-biasing.
9.  Excel import accepts the same date formats as CSV (not ISO-only).

Run with:
    uv run pytest tests/test_wave2_data.py -q
"""

from __future__ import annotations

import io
from datetime import date

import pytest
from httpx import AsyncClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.asset import Asset
from app.models.corporate_action import CorporateAction
from app.models.dividend import Dividend
from app.models.goal import Goal
from app.models.holding import Holding
from app.models.portfolio import Portfolio
from app.models.tax_record import TaxRecord
from app.models.transaction import Transaction
from app.models.user import User
from app.services.backup_service import export_portfolio_json, import_portfolio_json
from app.services.corporate_actions_service import apply_corporate_action
from app.services.csv_import_service import import_dividends
from app.services.esg_service import _weighted_average
from app.services.excel_service import parse_excel
from app.services.export_service import (
    export_workbook_xlsx,
    generate_portfolio_report_html,
)
from app.services.ofx_qif_import_service import import_statement, parse_ofx, parse_qif
from app.services.portfolio_service import calculate_cumulative_holding
from app.services.sheets_export_service import generate_portfolio_csv

# ---------------------------------------------------------------------------
# ORM helpers
# ---------------------------------------------------------------------------


async def _make_user(db: AsyncSession, email: str) -> User:
    user = User(email=email, password_hash="x", display_name="Wave2")
    db.add(user)
    await db.flush()
    return user


async def _make_portfolio(db: AsyncSession, user: User, name: str = "P") -> Portfolio:
    portfolio = Portfolio(user_id=user.id, name=name, currency="INR")
    db.add(portfolio)
    await db.flush()
    return portfolio


async def _make_holding(
    db: AsyncSession,
    portfolio: Portfolio,
    *,
    symbol: str = "RELIANCE",
    qty: float = 0.0,
    avg: float = 0.0,
    current_price: float | None = None,
) -> Holding:
    holding = Holding(
        portfolio_id=portfolio.id,
        stock_symbol=symbol,
        stock_name=symbol,
        exchange="NSE",
        currency="INR",
        cumulative_quantity=qty,
        average_price=avg,
        current_price=current_price,
    )
    db.add(holding)
    await db.flush()
    return holding


async def _add_tx(
    db: AsyncSession,
    holding: Holding,
    *,
    tx_type: str = "BUY",
    qty: float,
    price: float,
    when: date,
) -> Transaction:
    tx = Transaction(
        holding_id=holding.id,
        transaction_type=tx_type,
        date=when,
        quantity=qty,
        price=price,
        brokerage=0,
        source="MANUAL",
    )
    db.add(tx)
    await db.flush()
    return tx


# ===========================================================================
# 1a. QIF — one day/month convention per file
# ===========================================================================


QIF_DAYFIRST = """!Type:Invest
D13/04/2024
NBuy
YInfosys
I1500,50
Q10
^
D05/04/2024
NBuy
YInfosys
I1450.00
Q5
^
"""

QIF_MONTHFIRST = """!Type:Invest
D04/13/2024
NBuy
YInfosys
I1500.00
Q10
^
D04/05/2024
NBuy
YInfosys
I1450.00
Q5
^
"""

QIF_AMBIGUOUS = """!Type:Invest
D05/04/2024
NBuy
YInfosys
I1450.00
Q5
^
"""


class TestQifDateConvention:
    def test_dayfirst_file_parses_consistently(self):
        """13/04 proves day-first; 05/04 in the same file must follow it."""
        rows = parse_qif(QIF_DAYFIRST.encode())
        assert [r["date"] for r in rows] == [date(2024, 4, 13), date(2024, 4, 5)]

    def test_monthfirst_file_parses_consistently(self):
        """04/13 proves month-first; 04/05 must follow it (not become May 4)."""
        rows = parse_qif(QIF_MONTHFIRST.encode())
        assert [r["date"] for r in rows] == [date(2024, 4, 13), date(2024, 4, 5)]

    def test_ambiguous_file_defaults_dayfirst(self):
        """Fully ambiguous file → day-first (Indian/German target market)."""
        rows = parse_qif(QIF_AMBIGUOUS.encode())
        assert rows[0]["date"] == date(2024, 4, 5)

    def test_classic_apostrophe_two_digit_year(self):
        """Quicken "MM/DD'YY" still parses to the right century."""
        qif = "!Type:Invest\nD01/15'24\nNBuy\nYAcme\nI100.00\nQ1\n^\n"
        rows = parse_qif(qif.encode())
        assert rows[0]["date"] == date(2024, 1, 15)

    def test_qif_german_decimal_comma_price(self):
        rows = parse_qif(QIF_DAYFIRST.encode())
        assert rows[0]["price"] == pytest.approx(1500.50)


# ===========================================================================
# 1b. OFX — European decimal comma
# ===========================================================================


OFX_GERMAN = """OFXHEADER:100

<OFX>
<INVTRANLIST>
<BUYSTOCK>
<INVBUY>
<INVTRAN>
<DTTRADE>20240115
</INVTRAN>
<SECID>
<UNIQUEID>SAPX1
</SECID>
<UNITS>10
<UNITPRICE>1234,56
<COMMISSION>1,50
</INVBUY>
</BUYSTOCK>
</INVTRANLIST>
</OFX>
"""


class TestOfxGermanNumbers:
    def test_decimal_comma_price(self):
        rows = parse_ofx(OFX_GERMAN.encode())
        assert len(rows) == 1
        # "1234,56" must be 1234.56, not 123456.0 (naive comma stripping).
        assert rows[0]["price"] == pytest.approx(1234.56)
        assert rows[0]["brokerage"] == pytest.approx(1.50)
        assert rows[0]["quantity"] == pytest.approx(10.0)


# ===========================================================================
# 2. Corporate split — post-ex-date buys unaffected
# ===========================================================================


class TestSplitPostExDateBuys:
    async def test_split_adjusts_only_pre_ex_shares(self, db: AsyncSession):
        user = await _make_user(db, "split-postex@example.com")
        portfolio = await _make_portfolio(db, user)
        holding = await _make_holding(db, portfolio)

        # 10 shares before the ex-date, 10 bought after it.
        await _add_tx(db, holding, qty=10, price=100.0, when=date(2024, 1, 1))
        await _add_tx(db, holding, qty=10, price=50.0, when=date(2024, 6, 1))
        await calculate_cumulative_holding(holding.id, db)
        assert float(holding.cumulative_quantity) == pytest.approx(20.0)

        action = CorporateAction(
            holding_id=holding.id,
            action_type="SPLIT",
            ex_date=date(2024, 3, 1),
            ratio=2.0,
            status="DETECTED",
        )
        db.add(action)
        await db.flush()

        result = await apply_corporate_action(action.id, user.id, db)

        # Only the 10 pre-ex shares double (→ 20); the 10 post-ex shares are
        # untouched. Total 30 — NOT 40 (which the old current-quantity
        # multiplication produced).
        assert float(holding.cumulative_quantity) == pytest.approx(30.0)
        # Cost basis unchanged: 10*100 + 10*50 = 1500 → avg 50.
        assert float(holding.average_price) == pytest.approx(1500.0 / 30.0, abs=1e-4)
        assert result["details"]["applied"]["quantity_at_ex_date"] == pytest.approx(10.0)

        # Numbers derive from the ledger: a recompute reproduces them.
        await calculate_cumulative_holding(holding.id, db)
        assert float(holding.cumulative_quantity) == pytest.approx(30.0)

        # Idempotent: re-apply is a no-op.
        await apply_corporate_action(action.id, user.id, db)
        assert float(holding.cumulative_quantity) == pytest.approx(30.0)


# ===========================================================================
# 3. Dividend CSV re-import skips duplicates
# ===========================================================================


class TestDividendReimport:
    async def test_reimport_skips_existing(self, db: AsyncSession):
        user = await _make_user(db, "div-reimport@example.com")
        portfolio = await _make_portfolio(db, user)
        await _make_holding(db, portfolio, qty=10, avg=2500.0)

        rows = [{
            "stock_symbol": "RELIANCE",
            "exchange": "NSE",
            "ex_date": date(2024, 6, 15),
            "payment_date": date(2024, 7, 1),
            "amount_per_share": 10.5,
            "total_amount": 105.0,
            "is_reinvested": False,
            "reinvest_price": None,
            "reinvest_shares": None,
        }]

        first = await import_dividends(rows, portfolio.id, db)
        assert first["dividends_created"] == 1
        assert first["dividends_skipped"] == 0

        second = await import_dividends(rows, portfolio.id, db)
        assert second["dividends_created"] == 0
        assert second["dividends_skipped"] == 1

        count = len((await db.execute(select(Dividend))).scalars().all())
        assert count == 1


# ===========================================================================
# 4. Backup restore does not multiply user-level data
# ===========================================================================


class TestBackupRestoreDedup:
    async def test_double_restore_no_user_level_duplicates(self, db: AsyncSession):
        user = await _make_user(db, "backup-dedup@example.com")
        portfolio = await _make_portfolio(db, user, name="Main")
        holding = await _make_holding(db, portfolio, qty=10, avg=2500.0)
        await _add_tx(db, holding, qty=10, price=2500.0, when=date(2024, 1, 15))

        db.add(Goal(
            user_id=user.id, name="Retirement", target_amount=1_000_000.0,
            current_amount=0.0, category="RETIREMENT",
        ))
        db.add(Asset(user_id=user.id, asset_type="GOLD", name="Sovereign Bond"))
        db.add(TaxRecord(
            user_id=user.id, financial_year="2024-25", tax_jurisdiction="IN",
            gain_type="STCG", purchase_date=date(2024, 1, 15),
            sale_date=date(2024, 6, 20), purchase_price=25000.0,
            sale_price=30000.0, gain_amount=5000.0, tax_amount=1000.0,
            currency="INR",
        ))
        await db.flush()

        backup = await export_portfolio_json(portfolio.id, user.id, db)
        assert len(backup["goals"]) == 1
        assert len(backup["assets"]) == 1
        assert len(backup["tax_records"]) == 1

        # Restoring into the SAME user must not re-insert user-level data —
        # and restoring twice certainly must not.
        first = await import_portfolio_json(backup, user.id, db)
        second = await import_portfolio_json(backup, user.id, db)

        for summary in (first, second):
            assert summary["goals"] == 0
            assert summary["assets"] == 0
            assert summary["tax_records"] == 0
            assert summary["goals_skipped"] == 1
            assert summary["assets_skipped"] == 1
            assert summary["tax_records_skipped"] == 1
            # Portfolio-scoped data is still restored into the new portfolio.
            assert summary["holdings"] == 1
            assert summary["transactions"] == 1

        goals = (await db.execute(
            select(Goal).where(Goal.user_id == user.id)
        )).scalars().all()
        assets = (await db.execute(
            select(Asset).where(Asset.user_id == user.id)
        )).scalars().all()
        taxes = (await db.execute(
            select(TaxRecord).where(TaxRecord.user_id == user.id)
        )).scalars().all()
        assert len(goals) == 1
        assert len(assets) == 1
        assert len(taxes) == 1

    async def test_restore_into_fresh_user_creates_user_level_data(
        self, db: AsyncSession
    ):
        user = await _make_user(db, "backup-src@example.com")
        portfolio = await _make_portfolio(db, user, name="Src")
        db.add(Goal(
            user_id=user.id, name="House", target_amount=500_000.0,
            current_amount=0.0, category="HOUSE",
        ))
        await db.flush()

        backup = await export_portfolio_json(portfolio.id, user.id, db)

        other = await _make_user(db, "backup-dst@example.com")
        summary = await import_portfolio_json(backup, other.id, db)
        assert summary["goals"] == 1
        assert summary["goals_skipped"] == 0


# ===========================================================================
# 5. Unpriced holdings — "—", not -100 %
# ===========================================================================


class TestUnpricedHoldingExports:
    async def _portfolio_with_mixed_pricing(self, db: AsyncSession) -> Portfolio:
        user = await _make_user(db, "unpriced@example.com")
        portfolio = await _make_portfolio(db, user, name="Mixed")
        await _make_holding(
            db, portfolio, symbol="PRICED", qty=10, avg=2500.0, current_price=2800.0
        )
        await _make_holding(
            db, portfolio, symbol="NOPRICE", qty=10, avg=2500.0, current_price=None
        )
        return portfolio

    async def test_report_html_renders_dash_not_minus_100(self, db: AsyncSession):
        portfolio = await self._portfolio_with_mixed_pricing(db)
        html = await generate_portfolio_report_html(portfolio.id, "Tester", db)

        assert "-100" not in html
        assert ">—</td>" in html  # unpriced current/value/P&L cells
        assert "awaiting price data" in html
        # Totals exclude the unpriced holding: current value is 10 x 2800.
        assert "28,000.00" in html

    async def test_xlsx_workbook_dash_and_note(self, db: AsyncSession):
        portfolio = await self._portfolio_with_mixed_pricing(db)
        data = await export_workbook_xlsx(portfolio.id, db)

        wb = load_workbook(io.BytesIO(data))
        ws = wb["Holdings"]
        rows = {row[0].value: [c.value for c in row] for row in ws.iter_rows(min_row=2)}
        assert rows["NOPRICE"][4] == "—"  # Current Price
        assert rows["NOPRICE"][6] == "—"  # Current Value
        assert rows["NOPRICE"][7] == "—"  # P&L
        assert rows["NOPRICE"][8] == "—"  # P&L %
        assert -100 not in [v for v in rows["NOPRICE"] if isinstance(v, int | float)]

        summary = {row[0].value: row[1].value for row in wb["Summary"].iter_rows(min_row=2)}
        assert summary["Current Value"] == pytest.approx(28000.0)
        assert "awaiting price data" in str(summary.get("Note", ""))

    async def test_sheets_csv_dash_and_excluded_totals(self, db: AsyncSession):
        portfolio = await self._portfolio_with_mixed_pricing(db)
        csv_text = await generate_portfolio_csv(portfolio.id, db)

        assert "-100" not in csv_text
        noprice_line = next(
            line for line in csv_text.splitlines() if line.startswith("NOPRICE")
        )
        assert "—" in noprice_line
        # Totals row: market value only from the priced holding.
        total_line = next(
            line for line in csv_text.splitlines() if line.startswith("TOTAL")
        )
        assert "28000.0" in total_line


# ===========================================================================
# 6. Import provenance
# ===========================================================================


class TestImportProvenance:
    async def test_csv_route_stamps_source_csv(
        self, client: AsyncClient, auth_headers: dict[str, str]
    ):
        resp = await client.post(
            "/api/v1/portfolios/",
            json={"name": "Prov", "currency": "INR"},
            headers=auth_headers,
        )
        assert resp.status_code == 201
        pid = resp.json()["id"]

        csv_content = (
            "stock_symbol,stock_name,exchange,transaction_type,date,quantity,price\n"
            "RELIANCE,Reliance Industries,NSE,BUY,2024-01-15,10,2500\n"
        )
        resp = await client.post(
            f"/api/v1/import-export/csv?portfolio_id={pid}",
            files={"file": ("holdings.csv", csv_content.encode(), "text/csv")},
            headers=auth_headers,
        )
        assert resp.status_code == 200
        assert resp.json()["transactions_created"] == 1

        resp = await client.get(
            f"/api/v1/import-export/export/csv/{pid}/transactions",
            headers=auth_headers,
        )
        assert resp.status_code == 200
        reliance_rows = [
            line for line in resp.text.splitlines() if line.startswith("RELIANCE")
        ]
        assert reliance_rows
        assert ",CSV," in reliance_rows[0]
        assert ",EXCEL," not in reliance_rows[0]

    async def test_qif_import_stamps_source_qif(self, db: AsyncSession):
        user = await _make_user(db, "qif-prov@example.com")
        portfolio = await _make_portfolio(db, user)

        rows = parse_qif(QIF_DAYFIRST.encode())
        summary = await import_statement(rows, portfolio.id, db, source="QIF")
        assert summary["transactions_created"] == len(rows)

        txs = (await db.execute(select(Transaction))).scalars().all()
        assert txs
        assert all(tx.source == "QIF" for tx in txs)


# ===========================================================================
# 8. ESG weighted average — None sub-scores are skipped, not zeroed
# ===========================================================================


class TestEsgWeightedAverage:
    def test_none_values_are_skipped_not_zeroed(self):
        # Two equal-weight holdings; one is missing the governance score.
        # Zero-biasing would report 10.0 — the correct value is 20.0.
        assert _weighted_average([(20.0, 100.0), (None, 100.0)]) == pytest.approx(20.0)

    def test_weights_apply_only_to_present_values(self):
        assert _weighted_average(
            [(10.0, 100.0), (30.0, 300.0), (None, 600.0)]
        ) == pytest.approx((10.0 * 100 + 30.0 * 300) / 400)

    def test_all_none_returns_none(self):
        assert _weighted_average([(None, 100.0), (None, 50.0)]) is None

    def test_zero_weight_returns_none(self):
        assert _weighted_average([(20.0, 0.0)]) is None

    def test_empty_returns_none(self):
        assert _weighted_average([]) is None


# ===========================================================================
# 9. Excel import — shared date parsing (not ISO-only)
# ===========================================================================


def _excel_bytes(rows: list[list[object]]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.append([
        "stock_symbol", "stock_name", "exchange", "transaction_type",
        "date", "quantity", "price",
    ])
    for row in rows:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestExcelDateParsing:
    def test_accepts_dd_mm_yyyy_like_csv(self):
        data = _excel_bytes([
            ["RELIANCE", "Reliance", "NSE", "BUY", "15/01/2024", 10, 2500.0],
        ])
        rows = parse_excel(data)
        assert len(rows) == 1
        assert rows[0]["date"] == date(2024, 1, 15)

    def test_iso_still_works(self):
        data = _excel_bytes([
            ["RELIANCE", "Reliance", "NSE", "BUY", "2024-01-15", 10, 2500.0],
        ])
        rows = parse_excel(data)
        assert rows[0]["date"] == date(2024, 1, 15)

    def test_zero_price_row_kept(self):
        """The _missing() zero-guard: a legit 0-price allotment is not dropped."""
        data = _excel_bytes([
            ["RELIANCE", "Reliance", "NSE", "BUY", "2024-01-15", 10, 0],
        ])
        rows = parse_excel(data)
        assert len(rows) == 1
        assert rows[0]["price"] == 0.0
