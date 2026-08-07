"""Excel import / export service using openpyxl."""

from __future__ import annotations

import io
import logging
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.holding import Holding
from app.models.portfolio import Portfolio
from app.models.transaction import Transaction
from app.services.portfolio_service import calculate_cumulative_holding
from app.utils.dates import parse_date

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template column definitions
# ---------------------------------------------------------------------------

_TEMPLATE_COLUMNS = [
    "stock_symbol",
    "stock_name",
    "exchange",
    "transaction_type",  # BUY / SELL
    "date",              # YYYY-MM-DD
    "quantity",
    "price",
    "brokerage",
    "lower_mid_range_1",
    "lower_mid_range_2",
    "upper_mid_range_1",
    "upper_mid_range_2",
    "base_level",
    "top_level",
    "sector",
    "notes",
]


# ---------------------------------------------------------------------------
# Parse uploaded Excel
# ---------------------------------------------------------------------------

def parse_excel(file_bytes: bytes) -> list[dict]:
    """Parse an uploaded Excel file and return a list of structured row dicts.

    Expected columns match ``_TEMPLATE_COLUMNS``.
    Rows with missing required fields (symbol, name, exchange, type, date,
    quantity, price) are skipped with a warning.
    """
    wb = load_workbook(filename=io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    if ws is None:
        return []

    rows_iter = ws.iter_rows(values_only=False)

    # Read header row
    header_row = next(rows_iter, None)
    if header_row is None:
        return []

    headers = [
        str(cell.value).strip().lower().replace(" ", "_") if cell.value else ""
        for cell in header_row
    ]

    parsed: list[dict] = []

    for row in rows_iter:
        values = [cell.value for cell in row]
        row_dict: dict = {}
        for header, value in zip(headers, values):
            if header:
                row_dict[header] = value

        # Validate required fields
        symbol = row_dict.get("stock_symbol")
        name = row_dict.get("stock_name")
        exchange = row_dict.get("exchange")
        tx_type = row_dict.get("transaction_type")
        tx_date = row_dict.get("date")
        # Raw spreadsheet cell values (numeric, str, or None). Typed as Any so the
        # float() coercion below reads cleanly; the None/non-numeric cases are
        # filtered by the _missing() guard and caught by the try/except.
        qty: Any = row_dict.get("quantity")
        price: Any = row_dict.get("price")

        # A field is "missing" only when it is None or a blank string. Do NOT
        # treat a legitimate numeric 0 (e.g. a bonus/IPO allotment at price 0)
        # as absent — plain truthiness would drop those valid rows.
        def _missing(v: object) -> bool:
            return v is None or (isinstance(v, str) and not v.strip())

        if any(
            _missing(v)
            for v in (symbol, name, exchange, tx_type, tx_date, qty, price)
        ):
            logger.warning("Skipping row with missing required fields: %s", row_dict)
            continue

        # Normalise types
        row_dict["stock_symbol"] = str(symbol).strip().upper()
        row_dict["stock_name"] = str(name).strip()
        row_dict["exchange"] = str(exchange).strip().upper()
        row_dict["transaction_type"] = str(tx_type).strip().upper()
        if row_dict["transaction_type"] not in ("BUY", "SELL"):
            logger.warning("Invalid transaction type '%s' in row, skipping", tx_type)
            continue

        # Parse date via the shared parser (handles datetime/date cells and the
        # same string formats as the CSV importer, not just ISO).
        parsed_date = parse_date(tx_date)
        if parsed_date is None:
            logger.warning("Invalid date '%s' in row, skipping", tx_date)
            continue
        row_dict["date"] = parsed_date

        try:
            row_dict["quantity"] = float(qty)
            row_dict["price"] = float(price)
            row_dict["brokerage"] = float(row_dict.get("brokerage") or 0)
        except (ValueError, TypeError):
            logger.warning("Non-numeric quantity/price/brokerage in row, skipping: %s", row_dict)
            continue

        # Optional numeric fields
        for field in (
            "lower_mid_range_1",
            "lower_mid_range_2",
            "upper_mid_range_1",
            "upper_mid_range_2",
            "base_level",
            "top_level",
        ):
            val = row_dict.get(field)
            try:
                row_dict[field] = float(val) if val is not None else None
            except (ValueError, TypeError):
                row_dict[field] = None

        row_dict["sector"] = str(row_dict.get("sector", "")).strip() or None
        row_dict["notes"] = str(row_dict.get("notes", "")).strip() or None

        parsed.append(row_dict)

    wb.close()
    return parsed


# ---------------------------------------------------------------------------
# Import parsed data into a portfolio
# ---------------------------------------------------------------------------

async def import_to_portfolio(
    parsed_data: list[dict],
    portfolio_id: int,
    db: AsyncSession,
    *,
    source: str = "EXCEL",
) -> dict:
    """Create holdings and transactions from parsed import data.

    If a holding with the same symbol + exchange already exists in the
    portfolio, transactions are appended to it. Otherwise a new holding
    is created.

    ``source`` is stamped on every created transaction so provenance reflects
    the actual upload format (EXCEL / CSV / OFX / QIF), not always "EXCEL".

    Returns a summary dict with counts.
    """
    holdings_created = 0
    transactions_created = 0
    transactions_skipped = 0

    # Pre-load existing holdings for the portfolio
    result = await db.execute(
        select(Holding).where(Holding.portfolio_id == portfolio_id)
    )
    existing_holdings = result.scalars().all()
    holding_map: dict[str, Holding] = {
        f"{h.stock_symbol}|{h.exchange}": h for h in existing_holdings
    }

    # Pre-load existing transaction "fingerprints" so re-importing the same
    # statement (OFX/QIF/CSV/Excel all funnel through here) does not double
    # count. A transaction is a duplicate when its holding, date, type,
    # quantity and price all match an existing one. We can't rely on a DB
    # unique constraint (no migration), so dedup at the application level.
    def _tx_fingerprint(holding_key: str, tx_date, tx_type, qty, price) -> tuple:
        return (
            holding_key,
            tx_date,
            str(tx_type).upper(),
            round(float(qty), 6),
            round(float(price), 4),
        )

    seen_fingerprints: set[tuple] = set()
    if existing_holdings:
        id_to_key = {h.id: f"{h.stock_symbol}|{h.exchange}" for h in existing_holdings}
        existing_tx = await db.execute(
            select(Transaction).where(
                Transaction.holding_id.in_(id_to_key.keys())
            )
        )
        for tx in existing_tx.scalars().all():
            hkey = id_to_key.get(tx.holding_id)
            if hkey is not None:
                seen_fingerprints.add(
                    _tx_fingerprint(
                        hkey, tx.date, tx.transaction_type, tx.quantity, tx.price
                    )
                )

    for row in parsed_data:
        key = f"{row['stock_symbol']}|{row['exchange']}"
        holding = holding_map.get(key)

        if holding is None:
            holding = Holding(
                portfolio_id=portfolio_id,
                stock_symbol=row["stock_symbol"],
                stock_name=row["stock_name"],
                exchange=row["exchange"],
                cumulative_quantity=0.0,
                average_price=0.0,
                lower_mid_range_1=row.get("lower_mid_range_1"),
                lower_mid_range_2=row.get("lower_mid_range_2"),
                upper_mid_range_1=row.get("upper_mid_range_1"),
                upper_mid_range_2=row.get("upper_mid_range_2"),
                base_level=row.get("base_level"),
                top_level=row.get("top_level"),
                sector=row.get("sector"),
                notes=row.get("notes"),
            )
            db.add(holding)
            await db.flush()
            await db.refresh(holding)
            holding_map[key] = holding
            holdings_created += 1
        else:
            # Update range levels from Excel if provided (non-None)
            for field in (
                "lower_mid_range_1",
                "lower_mid_range_2",
                "upper_mid_range_1",
                "upper_mid_range_2",
                "base_level",
                "top_level",
            ):
                val = row.get(field)
                if val is not None:
                    setattr(holding, field, val)

        # Skip duplicates (same holding + date + type + quantity + price).
        fingerprint = _tx_fingerprint(
            key,
            row["date"],
            row["transaction_type"],
            row["quantity"],
            row["price"],
        )
        if fingerprint in seen_fingerprints:
            transactions_skipped += 1
            continue
        seen_fingerprints.add(fingerprint)

        # Create the transaction
        tx = Transaction(
            holding_id=holding.id,
            transaction_type=row["transaction_type"],
            date=row["date"],
            quantity=row["quantity"],
            price=row["price"],
            brokerage=row.get("brokerage", 0),
            notes=row.get("notes"),
            source=source,
        )
        db.add(tx)
        transactions_created += 1

    await db.flush()

    # Recalculate all affected holdings
    recalculated = set()
    for row in parsed_data:
        key = f"{row['stock_symbol']}|{row['exchange']}"
        holding = holding_map.get(key)
        if holding and holding.id not in recalculated:
            await calculate_cumulative_holding(holding.id, db)
            recalculated.add(holding.id)

    return {
        "holdings_created": holdings_created,
        "transactions_created": transactions_created,
        "transactions_skipped": transactions_skipped,
        "holdings_updated": len(recalculated) - holdings_created,
    }


# ---------------------------------------------------------------------------
# Export portfolio to Excel bytes
# ---------------------------------------------------------------------------

async def export_portfolio(portfolio_id: int, db: AsyncSession) -> bytes:
    """Generate an Excel workbook with all holdings and transactions for a portfolio.

    Returns the workbook as bytes suitable for a StreamingResponse.
    """
    result = await db.execute(
        select(Portfolio)
        .options(selectinload(Portfolio.holdings).selectinload(Holding.transactions))
        .where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if portfolio is None:
        raise ValueError(f"Portfolio {portfolio_id} not found")

    wb = Workbook()

    # ── Holdings sheet ──────────────────────────────────────────────────
    ws_holdings = wb.active
    ws_holdings.title = "Holdings"

    holding_headers = [
        "Stock Symbol",
        "Stock Name",
        "Exchange",
        "Quantity",
        "Avg Price",
        "Current Price",
        "Action",
        "RSI",
        "Sector",
        "Lower Mid 1",
        "Lower Mid 2",
        "Upper Mid 1",
        "Upper Mid 2",
        "Base Level",
        "Top Level",
        "Notes",
    ]
    _write_header(ws_holdings, holding_headers)

    for h in portfolio.holdings:
        ws_holdings.append(
            [
                h.stock_symbol,
                h.stock_name,
                h.exchange,
                float(h.cumulative_quantity),
                float(h.average_price),
                float(h.current_price) if h.current_price else None,
                h.action_needed,
                h.current_rsi,
                h.sector,
                float(h.lower_mid_range_1) if h.lower_mid_range_1 else None,
                float(h.lower_mid_range_2) if h.lower_mid_range_2 else None,
                float(h.upper_mid_range_1) if h.upper_mid_range_1 else None,
                float(h.upper_mid_range_2) if h.upper_mid_range_2 else None,
                float(h.base_level) if h.base_level else None,
                float(h.top_level) if h.top_level else None,
                h.notes,
            ]
        )

    # ── Transactions sheet ──────────────────────────────────────────────
    ws_txns = wb.create_sheet("Transactions")
    tx_headers = [
        "Stock Symbol",
        "Exchange",
        "Type",
        "Date",
        "Quantity",
        "Price",
        "Brokerage",
        "Source",
        "Notes",
    ]
    _write_header(ws_txns, tx_headers)

    for h in portfolio.holdings:
        for tx in h.transactions:
            ws_txns.append(
                [
                    h.stock_symbol,
                    h.exchange,
                    tx.transaction_type,
                    tx.date.isoformat(),
                    float(tx.quantity),
                    float(tx.price),
                    float(tx.brokerage),
                    tx.source,
                    tx.notes,
                ]
            )

    # Auto-size columns
    for ws in [ws_holdings, ws_txns]:
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Generate blank template
# ---------------------------------------------------------------------------

def generate_template() -> bytes:
    """Generate a blank Excel template with the expected column headers
    and a sample row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"

    _write_header(ws, _TEMPLATE_COLUMNS)

    # Sample row
    ws.append(
        [
            "RELIANCE",
            "Reliance Industries Ltd",
            "NSE",
            "BUY",
            "2024-01-15",
            10,
            2500.00,
            50.00,
            2300.00,
            2100.00,
            2700.00,
            2900.00,
            2000.00,
            3000.00,
            "Energy",
            "Initial purchase",
        ]
    )

    # Auto-size
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _write_header(ws, headers: list[str]) -> None:
    """Write a styled header row to a worksheet."""
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center")

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
