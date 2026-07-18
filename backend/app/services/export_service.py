"""Export services: CSV and PDF report generation."""

from __future__ import annotations

import csv
import html as html_mod
import io
import json
import logging
import zipfile
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.models.holding import Holding
from app.models.portfolio import Portfolio
from app.models.transaction import Transaction
from app.models.tax_record import TaxRecord
from app.services.backup_service import export_portfolio_json
from app.services.tax_service import generate_tax_summary

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _sanitize_csv_cell(value: object) -> object:
    """Prevent CSV formula injection by prefixing dangerous characters.

    Spreadsheet applications (Excel, Google Sheets) interpret cells starting
    with ``=``, ``+``, ``-``, ``@``, ``\\t``, or ``\\r`` as formulas or
    field separators.  Prefixing with a single quote neutralises this.
    """
    if isinstance(value, str) and value and value[0] in "=+-@\t\r":
        return f"'{value}"
    return value


# ---------------------------------------------------------------------------
# CSV Export
# ---------------------------------------------------------------------------


async def export_holdings_csv(portfolio_id: int, db: AsyncSession) -> str:
    """Export portfolio holdings as CSV string."""
    result = await db.execute(
        select(Portfolio)
        .options(selectinload(Portfolio.holdings))
        .where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if portfolio is None:
        raise ValueError(f"Portfolio {portfolio_id} not found")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Stock Symbol", "Stock Name", "Exchange", "Quantity", "Avg Price",
        "Current Price", "P&L %", "Action Needed", "RSI", "Sector",
    ])

    for h in portfolio.holdings:
        qty = float(h.cumulative_quantity)
        avg = float(h.average_price)
        current = float(h.current_price) if h.current_price else None
        pnl = None
        if current and avg > 0:
            pnl = round(((current - avg) / avg) * 100, 2)
        writer.writerow([
            _sanitize_csv_cell(h.stock_symbol),
            _sanitize_csv_cell(h.stock_name),
            _sanitize_csv_cell(h.exchange),
            qty, avg, current, pnl,
            _sanitize_csv_cell(h.action_needed),
            h.current_rsi,
            _sanitize_csv_cell(h.sector),
        ])

    return output.getvalue()


async def export_transactions_csv(portfolio_id: int, db: AsyncSession) -> str:
    """Export all transactions for a portfolio as CSV."""
    result = await db.execute(
        select(Portfolio)
        .options(selectinload(Portfolio.holdings).selectinload(Holding.transactions))
        .where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if portfolio is None:
        raise ValueError(f"Portfolio {portfolio_id} not found")

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "Stock Symbol", "Exchange", "Type", "Date", "Quantity",
        "Price", "Brokerage", "Source", "Notes",
    ])

    for h in portfolio.holdings:
        for tx in h.transactions:
            writer.writerow([
                _sanitize_csv_cell(h.stock_symbol),
                _sanitize_csv_cell(h.exchange),
                _sanitize_csv_cell(tx.transaction_type),
                tx.date.isoformat(),
                float(tx.quantity), float(tx.price), float(tx.brokerage),
                _sanitize_csv_cell(tx.source),
                _sanitize_csv_cell(tx.notes),
            ])

    return output.getvalue()


# ---------------------------------------------------------------------------
# PDF Report (simple HTML-based)
# ---------------------------------------------------------------------------


async def generate_portfolio_report_html(
    portfolio_id: int,
    user_name: str,
    db: AsyncSession,
) -> str:
    """Generate a portfolio summary report as styled HTML (can be printed to PDF)."""
    result = await db.execute(
        select(Portfolio)
        .options(selectinload(Portfolio.holdings))
        .where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if portfolio is None:
        raise ValueError(f"Portfolio {portfolio_id} not found")

    total_invested = 0.0
    total_current = 0.0
    holdings_data = []

    for h in portfolio.holdings:
        qty = float(h.cumulative_quantity)
        avg = float(h.average_price)
        invested = qty * avg
        current_price = float(h.current_price) if h.current_price else 0
        current_value = qty * current_price
        pnl = current_value - invested
        pnl_pct = (pnl / invested * 100) if invested > 0 else 0

        total_invested += invested
        total_current += current_value

        holdings_data.append({
            "symbol": h.stock_symbol,
            "name": h.stock_name,
            "exchange": h.exchange,
            "qty": qty,
            "avg": avg,
            "current": current_price,
            "invested": invested,
            "value": current_value,
            "pnl": pnl,
            "pnl_pct": pnl_pct,
            "action": h.action_needed,
            "rsi": h.current_rsi,
        })

    total_pnl = total_current - total_invested
    total_pnl_pct = (total_pnl / total_invested * 100) if total_invested > 0 else 0

    now = datetime.now().strftime("%B %d, %Y")

    # Build HTML report (escape user-controlled strings to prevent XSS)
    _esc = html_mod.escape
    rows_html = ""
    for h in holdings_data:
        pnl_color = "#16a34a" if h["pnl"] >= 0 else "#dc2626"
        rows_html += f"""
        <tr>
            <td>{_esc(str(h['symbol']))}</td>
            <td>{_esc(str(h['name']))}</td>
            <td style="text-align:right">{h['qty']:.2f}</td>
            <td style="text-align:right">{h['avg']:.2f}</td>
            <td style="text-align:right">{h['current']:.2f}</td>
            <td style="text-align:right">{h['invested']:,.2f}</td>
            <td style="text-align:right">{h['value']:,.2f}</td>
            <td style="text-align:right;color:{pnl_color}">{h['pnl']:+,.2f} ({h['pnl_pct']:+.1f}%)</td>
            <td style="text-align:center">{_esc(str(h['action']))}</td>
            <td style="text-align:right">{f"{h['rsi']:.1f}" if h['rsi'] else '--'}</td>
        </tr>"""

    total_pnl_color = "#16a34a" if total_pnl >= 0 else "#dc2626"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Portfolio Report - {_esc(portfolio.name)}</title>
<style>
    body {{ font-family: -apple-system, 'Segoe UI', sans-serif; margin: 40px; color: #1a1a1a; }}
    h1 {{ color: #1e3a5f; margin-bottom: 4px; }}
    .subtitle {{ color: #666; margin-bottom: 24px; }}
    .summary {{ display: flex; gap: 24px; margin-bottom: 32px; }}
    .summary-card {{ background: #f8f9fa; border-radius: 8px; padding: 16px 24px; flex: 1; }}
    .summary-card .label {{ font-size: 12px; color: #666; text-transform: uppercase; }}
    .summary-card .value {{ font-size: 24px; font-weight: 700; margin-top: 4px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th {{ background: #1e3a5f; color: white; padding: 10px 8px; text-align: left; }}
    td {{ padding: 8px; border-bottom: 1px solid #e5e7eb; }}
    tr:hover {{ background: #f8f9fa; }}
    .footer {{ margin-top: 32px; font-size: 11px; color: #999; text-align: center; }}
</style></head><body>
<h1>{_esc(portfolio.name)} — Portfolio Report</h1>
<p class="subtitle">Generated on {now} for {_esc(user_name)} | Currency: {_esc(portfolio.currency)}</p>
<div class="summary">
    <div class="summary-card">
        <div class="label">Total Invested</div>
        <div class="value">{total_invested:,.2f}</div>
    </div>
    <div class="summary-card">
        <div class="label">Current Value</div>
        <div class="value">{total_current:,.2f}</div>
    </div>
    <div class="summary-card">
        <div class="label">Total P&L</div>
        <div class="value" style="color:{total_pnl_color}">{total_pnl:+,.2f} ({total_pnl_pct:+.1f}%)</div>
    </div>
    <div class="summary-card">
        <div class="label">Holdings</div>
        <div class="value">{len(holdings_data)}</div>
    </div>
</div>
<table>
<thead><tr>
    <th>Symbol</th><th>Name</th><th style="text-align:right">Qty</th>
    <th style="text-align:right">Avg Price</th><th style="text-align:right">Current</th>
    <th style="text-align:right">Invested</th><th style="text-align:right">Value</th>
    <th style="text-align:right">P&L</th><th style="text-align:center">Action</th>
    <th style="text-align:right">RSI</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>
<div class="footer">FinanceTracker — Personal Investment Portfolio Tracking | Generated automatically</div>
</body></html>"""

    return html


# ---------------------------------------------------------------------------
# Capital-gains / ITR-ready tax report export
# ---------------------------------------------------------------------------


async def _load_tax_records(
    user_id: int,
    financial_year: str,
    jurisdiction: str,
    db: AsyncSession,
) -> list[TaxRecord]:
    """Load a user's TaxRecords for a FY + jurisdiction, eager-loading the
    linked transaction/holding so the stock symbol can be resolved.

    Scoped by ``user_id`` — records are only ever returned for the requesting
    user. Ordered by gain type then sale date for a stable, readable statement.
    """
    result = await db.execute(
        select(TaxRecord)
        .options(
            selectinload(TaxRecord.transaction).selectinload(Transaction.holding)
        )
        .where(
            TaxRecord.user_id == user_id,
            TaxRecord.financial_year == financial_year,
            TaxRecord.tax_jurisdiction == jurisdiction,
        )
        .order_by(TaxRecord.gain_type, TaxRecord.sale_date)
    )
    return list(result.scalars().all())


def _record_symbol(record: TaxRecord) -> str:
    """Resolve the stock symbol for a tax record via its linked transaction's
    holding, falling back to an empty string when the link is missing (e.g. the
    transaction was deleted and the FK set to NULL)."""
    txn = record.transaction
    if txn is not None and txn.holding is not None:
        return txn.holding.stock_symbol or ""
    return ""


def _record_quantity(record: TaxRecord) -> float | None:
    """Best-effort per-bucket quantity for a tax record.

    ``purchase_price`` / ``sale_price`` on a TaxRecord store the *aggregated*
    cost basis and proceeds for the consumed lots, not per-unit prices. The
    matched quantity is therefore ``proceeds / sale_unit_price``, where the
    per-unit sale price comes from the linked SELL transaction. Returns ``None``
    when it cannot be derived (no transaction, or a zero/absent unit price).
    """
    txn = record.transaction
    if txn is None or record.sale_price is None:
        return None
    unit_price = float(txn.price)
    if unit_price <= 0:
        return None
    return round(float(record.sale_price) / unit_price, 4)


def _num(value: object) -> float | None:
    """Coerce a Decimal/None DB numeric to float (or None)."""
    return float(value) if value is not None else None  # type: ignore[arg-type]


async def export_tax_report_csv(
    user_id: int,
    financial_year: str,
    jurisdiction: str,
    db: AsyncSession,
) -> str:
    """Build an ITR-ready capital-gains CSV for a user + FY + jurisdiction.

    Layout: a per-record table (one row per TaxRecord) followed by a totals
    summary section derived from ``generate_tax_summary``. All free-text /
    user-influenced cells are passed through ``_sanitize_csv_cell`` to defuse
    spreadsheet formula injection. An empty record set still produces a valid
    file with the header, a "no records" note, and a zeroed summary.
    """
    records = await _load_tax_records(user_id, financial_year, jurisdiction, db)
    summary = await generate_tax_summary(
        user_id=user_id,
        financial_year=financial_year,
        jurisdiction=jurisdiction,
        db=db,
    )

    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow([
        "Symbol", "Gain Type", "Purchase Date", "Sale Date",
        "Holding Period (Days)", "Quantity", "Cost Basis", "Proceeds",
        "Gain Amount", "Tax Amount", "Currency",
    ])

    if not records:
        writer.writerow([
            _sanitize_csv_cell(
                f"No tax records for {financial_year} ({jurisdiction})"
            )
        ])
    else:
        for r in records:
            qty = _record_quantity(r)
            writer.writerow([
                _sanitize_csv_cell(_record_symbol(r)),
                _sanitize_csv_cell(r.gain_type),
                r.purchase_date.isoformat() if r.purchase_date else "",
                r.sale_date.isoformat() if r.sale_date else "",
                r.holding_period_days if r.holding_period_days is not None else "",
                qty if qty is not None else "",
                _num(r.purchase_price),
                _num(r.sale_price),
                _num(r.gain_amount),
                _num(r.tax_amount),
                _sanitize_csv_cell(r.currency),
            ])

    # ── Totals summary section ─────────────────────────────────────────
    writer.writerow([])
    writer.writerow([_sanitize_csv_cell("SUMMARY")])
    writer.writerow(["Financial Year", _sanitize_csv_cell(summary["financial_year"])])
    writer.writerow(["Jurisdiction", _sanitize_csv_cell(summary["tax_jurisdiction"])])
    writer.writerow(["Total STCG", summary["total_stcg"]])
    writer.writerow(["Total LTCG", summary["total_ltcg"]])
    writer.writerow(["Total Tax", summary["total_tax"]])
    writer.writerow(["Exemption Used", summary["exemption_used"]])
    writer.writerow(["Records Count", summary["records_count"]])

    return output.getvalue()


async def generate_tax_report_html(
    user_id: int,
    user_name: str,
    financial_year: str,
    jurisdiction: str,
    db: AsyncSession,
) -> str:
    """Generate a self-contained, printable capital-gains statement as HTML.

    Mirrors the styling/escaping of ``generate_portfolio_report_html``: a titled
    statement with a per-record table and a summary block (total STCG, total
    LTCG, total tax, exemption used). An empty record set produces a valid page
    noting that no records exist for the period.
    """
    records = await _load_tax_records(user_id, financial_year, jurisdiction, db)
    summary = await generate_tax_summary(
        user_id=user_id,
        financial_year=financial_year,
        jurisdiction=jurisdiction,
        db=db,
    )

    _esc = html_mod.escape
    now = datetime.now().strftime("%B %d, %Y")

    def _fmt(value: object) -> str:
        num = _num(value)
        return f"{num:,.2f}" if num is not None else "--"

    if records:
        rows_html = ""
        for r in records:
            gain = _num(r.gain_amount)
            gain_color = "#16a34a" if (gain is None or gain >= 0) else "#dc2626"
            qty = _record_quantity(r)
            rows_html += f"""
        <tr>
            <td>{_esc(_record_symbol(r) or '--')}</td>
            <td>{_esc(str(r.gain_type))}</td>
            <td>{r.purchase_date.isoformat() if r.purchase_date else '--'}</td>
            <td>{r.sale_date.isoformat() if r.sale_date else '--'}</td>
            <td style="text-align:right">{r.holding_period_days if r.holding_period_days is not None else '--'}</td>
            <td style="text-align:right">{f'{qty:,.4f}' if qty is not None else '--'}</td>
            <td style="text-align:right">{_fmt(r.purchase_price)}</td>
            <td style="text-align:right">{_fmt(r.sale_price)}</td>
            <td style="text-align:right;color:{gain_color}">{f'{gain:+,.2f}' if gain is not None else '--'}</td>
            <td style="text-align:right">{_fmt(r.tax_amount)}</td>
            <td style="text-align:center">{_esc(str(r.currency))}</td>
        </tr>"""
        table_html = f"""<table>
<thead><tr>
    <th>Symbol</th><th>Gain Type</th><th>Purchase Date</th><th>Sale Date</th>
    <th style="text-align:right">Days Held</th><th style="text-align:right">Qty</th>
    <th style="text-align:right">Cost Basis</th><th style="text-align:right">Proceeds</th>
    <th style="text-align:right">Gain</th><th style="text-align:right">Tax</th>
    <th style="text-align:center">Currency</th>
</tr></thead>
<tbody>{rows_html}</tbody>
</table>"""
    else:
        table_html = (
            '<p class="empty">No capital-gains tax records found for '
            f"{_esc(financial_year)} ({_esc(jurisdiction)}).</p>"
        )

    stcg_color = "#16a34a" if summary["total_stcg"] >= 0 else "#dc2626"
    ltcg_color = "#16a34a" if summary["total_ltcg"] >= 0 else "#dc2626"

    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<title>Capital Gains Statement — {_esc(financial_year)} ({_esc(jurisdiction)})</title>
<style>
    body {{ font-family: -apple-system, 'Segoe UI', sans-serif; margin: 40px; color: #1a1a1a; }}
    h1 {{ color: #1e3a5f; margin-bottom: 4px; }}
    .subtitle {{ color: #666; margin-bottom: 24px; }}
    .summary {{ display: flex; gap: 24px; margin-bottom: 32px; flex-wrap: wrap; }}
    .summary-card {{ background: #f8f9fa; border-radius: 8px; padding: 16px 24px; flex: 1; min-width: 160px; }}
    .summary-card .label {{ font-size: 12px; color: #666; text-transform: uppercase; }}
    .summary-card .value {{ font-size: 24px; font-weight: 700; margin-top: 4px; }}
    table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
    th {{ background: #1e3a5f; color: white; padding: 10px 8px; text-align: left; }}
    td {{ padding: 8px; border-bottom: 1px solid #e5e7eb; }}
    tr:hover {{ background: #f8f9fa; }}
    .empty {{ color: #666; font-style: italic; padding: 24px 0; }}
    .footer {{ margin-top: 32px; font-size: 11px; color: #999; text-align: center; }}
</style></head><body>
<h1>Capital Gains Statement</h1>
<p class="subtitle">Financial Year {_esc(financial_year)} · Jurisdiction {_esc(jurisdiction)} · Generated on {now} for {_esc(user_name)}</p>
<div class="summary">
    <div class="summary-card">
        <div class="label">Total STCG</div>
        <div class="value" style="color:{stcg_color}">{summary['total_stcg']:,.2f}</div>
    </div>
    <div class="summary-card">
        <div class="label">Total LTCG</div>
        <div class="value" style="color:{ltcg_color}">{summary['total_ltcg']:,.2f}</div>
    </div>
    <div class="summary-card">
        <div class="label">Total Tax</div>
        <div class="value">{summary['total_tax']:,.2f}</div>
    </div>
    <div class="summary-card">
        <div class="label">Exemption Used</div>
        <div class="value">{summary['exemption_used']:,.2f}</div>
    </div>
    <div class="summary-card">
        <div class="label">Records</div>
        <div class="value">{summary['records_count']}</div>
    </div>
</div>
{table_html}
<div class="footer">FinanceTracker — Capital Gains Statement | Generated automatically. Verify figures against broker statements before filing.</div>
</body></html>"""

    return html


# ---------------------------------------------------------------------------
# PDF Export (requires xhtml2pdf)
# ---------------------------------------------------------------------------


async def generate_portfolio_pdf(
    portfolio_id: int, user_name: str, db: AsyncSession
) -> bytes:
    """Generate a PDF report from the HTML portfolio report.

    Requires the ``xhtml2pdf`` package. Raises ImportError if not installed.
    """
    from xhtml2pdf import pisa  # noqa: F811

    html_content = await generate_portfolio_report_html(portfolio_id, user_name, db)
    output = io.BytesIO()
    pisa.CreatePDF(html_content, dest=output)
    return output.getvalue()


# ---------------------------------------------------------------------------
# Multi-sheet XLSX workbook (Holdings + Transactions + Dividends + Summary)
# ---------------------------------------------------------------------------


def _xlsx_header(ws, headers: list[str]) -> None:
    """Write a bold, filled header row to a worksheet."""
    font = Font(bold=True, color="FFFFFF")
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    align = Alignment(horizontal="center")
    for idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = align


def _xlsx_finalize(ws) -> None:
    """Freeze the header row and auto-size columns for a worksheet."""
    ws.freeze_panes = "A2"
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[letter].width = min(max_len + 2, 40)


async def export_workbook_xlsx(portfolio_id: int, db: AsyncSession) -> bytes:
    """Build a multi-sheet XLSX workbook for a portfolio.

    Sheets: Holdings, Transactions, Dividends, Summary. Raises ``ValueError``
    if the portfolio does not exist or has no holdings (the API layer maps this
    to a 404, mirroring ``excel_service.export_portfolio``).
    """
    result = await db.execute(
        select(Portfolio)
        .options(
            selectinload(Portfolio.holdings).selectinload(Holding.transactions),
            selectinload(Portfolio.holdings).selectinload(Holding.dividends),
        )
        .where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if portfolio is None:
        raise ValueError(f"Portfolio {portfolio_id} not found")
    if not portfolio.holdings:
        raise ValueError(f"Portfolio {portfolio_id} has no holdings to export")

    wb = Workbook()

    # ── Holdings ────────────────────────────────────────────────────────
    ws_h = wb.active
    ws_h.title = "Holdings"
    _xlsx_header(ws_h, [
        "Symbol", "Exchange", "Quantity", "Avg Price", "Current Price",
        "Invested", "Current Value", "P&L", "P&L %",
    ])

    total_invested = 0.0
    total_current = 0.0
    for h in portfolio.holdings:
        qty = float(h.cumulative_quantity)
        avg = float(h.average_price)
        cur = float(h.current_price) if h.current_price else 0.0
        invested = qty * avg
        current_value = qty * cur
        pnl = current_value - invested
        pnl_pct = (pnl / invested * 100) if invested > 0 else 0.0
        total_invested += invested
        total_current += current_value
        ws_h.append([
            h.stock_symbol, h.exchange, qty, avg, cur or None,
            round(invested, 2), round(current_value, 2),
            round(pnl, 2), round(pnl_pct, 2),
        ])

    # ── Transactions ────────────────────────────────────────────────────
    ws_t = wb.create_sheet("Transactions")
    _xlsx_header(ws_t, [
        "Date", "Symbol", "Type", "Quantity", "Price", "Total", "Fees", "Notes",
    ])
    for h in portfolio.holdings:
        for tx in h.transactions:
            qty = float(tx.quantity)
            price = float(tx.price)
            ws_t.append([
                tx.date.isoformat(), h.stock_symbol, tx.transaction_type,
                qty, price, round(qty * price, 2), float(tx.brokerage), tx.notes,
            ])

    # ── Dividends ───────────────────────────────────────────────────────
    ws_d = wb.create_sheet("Dividends")
    _xlsx_header(ws_d, ["Ex-Date", "Symbol", "Amount", "Currency", "DRIP?"])
    for h in portfolio.holdings:
        for d in h.dividends:
            ws_d.append([
                d.ex_date.isoformat(), h.stock_symbol, float(d.total_amount),
                h.currency, "Yes" if d.is_reinvested else "No",
            ])

    # ── Summary ─────────────────────────────────────────────────────────
    ws_s = wb.create_sheet("Summary")
    _xlsx_header(ws_s, ["Metric", "Value"])
    total_pnl = total_current - total_invested
    total_pnl_pct = (total_pnl / total_invested * 100) if total_invested > 0 else 0.0
    for label, value in [
        ("Portfolio", portfolio.name),
        ("Currency", portfolio.currency),
        ("Total Invested", round(total_invested, 2)),
        ("Current Value", round(total_current, 2)),
        ("Total P&L", round(total_pnl, 2)),
        ("Total P&L %", round(total_pnl_pct, 2)),
        ("Holdings Count", len(portfolio.holdings)),
        ("Generated At", datetime.now().isoformat(timespec="seconds")),
    ]:
        ws_s.append([label, value])

    for ws in (ws_h, ws_t, ws_d, ws_s):
        _xlsx_finalize(ws)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ---------------------------------------------------------------------------
# "Export everything" ZIP bundle
# ---------------------------------------------------------------------------


async def export_everything_zip(
    portfolio_id: int, user_name: str, db: AsyncSession
) -> bytes:
    """Build an in-memory ZIP with CSVs, JSON backup, HTML report, the XLSX
    workbook, and (best-effort) the PDF report.

    Raises ``ValueError`` if the portfolio does not exist / has no holdings.
    A missing ``xhtml2pdf`` (or any PDF-generation failure) simply omits the
    PDF instead of failing the whole bundle.
    """
    result = await db.execute(
        select(Portfolio).where(Portfolio.id == portfolio_id)
    )
    portfolio = result.scalar_one_or_none()
    if portfolio is None:
        raise ValueError(f"Portfolio {portfolio_id} not found")

    holdings_csv = await export_holdings_csv(portfolio_id, db)
    transactions_csv = await export_transactions_csv(portfolio_id, db)
    backup = await export_portfolio_json(portfolio_id, portfolio.user_id, db)
    report_html = await generate_portfolio_report_html(portfolio_id, user_name, db)
    workbook = await export_workbook_xlsx(portfolio_id, db)

    pdf_bytes: bytes | None = None
    try:
        pdf_bytes = await generate_portfolio_pdf(portfolio_id, user_name, db)
    except ImportError:
        logger.info("xhtml2pdf not installed; omitting PDF from export bundle")
    except Exception:  # pragma: no cover - defensive; never fail the bundle
        logger.warning("PDF generation failed for export bundle", exc_info=True)

    ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    contents = [
        "holdings.csv",
        "transactions.csv",
        "portfolio_backup.json",
        "report.html",
        "portfolio_workbook.xlsx",
    ]

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("holdings.csv", holdings_csv)
        zf.writestr("transactions.csv", transactions_csv)
        zf.writestr(
            "portfolio_backup.json",
            json.dumps(backup, indent=2, ensure_ascii=False),
        )
        zf.writestr("report.html", report_html)
        zf.writestr("portfolio_workbook.xlsx", workbook)
        if pdf_bytes:
            zf.writestr("portfolio_report.pdf", pdf_bytes)
            contents.append("portfolio_report.pdf")

        readme = (
            "FinanceTracker export bundle\n"
            f"Portfolio: {portfolio.name}\n"
            f"Generated: {ts}\n\n"
            "Contents:\n" + "\n".join(f"  - {c}" for c in contents) + "\n"
        )
        if not pdf_bytes:
            readme += (
                "\nNote: PDF report omitted "
                "(xhtml2pdf not installed or generation failed).\n"
            )
        zf.writestr("README.txt", readme)

    buf.seek(0)
    return buf.read()
